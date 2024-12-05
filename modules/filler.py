import json
from pathlib import Path
from typing import IO, Any, Dict, List, OrderedDict

from docxtpl import DocxTemplate
from openpyxl import load_workbook
from pydantic import BaseModel
from PyPDFForm import FormWrapper, PdfWrapper

# from modules.xlsx_image_loader import SheetImageLoader


class FormFiller(BaseModel):
    template_filename: str
    output_filename: str
    fields: OrderedDict

def print_pdf_widgets(path):
    pdf_form_schema = PdfWrapper(path).schema
    print(json.dumps(pdf_form_schema, indent=4, sort_keys=True))
    print("")
    for k, v in pdf_form_schema["properties"].items():
        print(f"{k} : {v['type']}")
        # print(f"{k}")
        pass


def get_form_fillers(data_sheet="assets/fiche.xlsx") -> List[FormFiller]:
    wb = load_workbook(filename=data_sheet, data_only=True)
    ws = wb["Suivi"]
    
    # try:
    #     image_loader = SheetImageLoader(ws)
    #     anchors = image_loader.get_image_anchors()
    #     if len(anchors) > 1:
    #         signature_image = image_loader.get(anchors[0])
    # except:
    #     print("Erreur lors du chargement de l'image")
    #     pass
    
    form_fillers: List[FormFiller] = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        if (
            ws["A1"].value == "TEMPLATE FILENAME"
            and ws["A2"].value == "OUTPUT FILENAME"
        ):
            # the sheet represent a form filler
            template_filename = ws["B1"].value
            output_filename = ws["B2"].value

            fields = OrderedDict()
            for i in range(4, ws.max_row + 1):
                k = ws[f"B{i}"].value
                v = ws[f"C{i}"].value
                if k != None and k != "":
                    # if isinstance(v, int) or isinstance(v, float):
                    #     v = str(v)
                    fields[k] = v

            form_filler = FormFiller(
                template_filename=template_filename,
                output_filename=output_filename,
                fields=fields,
            )

            form_fillers.append(form_filler)

    return form_fillers


def fill_forms(
    form_fillers: List[FormFiller],
    editable=False,
    template_base_dir: str = "assets/templates/",
    output_base_dir: str = "tmp/",
) -> List[Path]:
    
    output_filepaths: List[Path] = []
    
    for form_filler in form_fillers:
        p = Path(template_base_dir).joinpath(form_filler.template_filename)
        if p.exists():
            template_filepath = str(p)

            output_filepath = str(
                Path(output_base_dir).joinpath(form_filler.output_filename)
            )
            
            fields = dict(form_filler.fields.items())

            if Path(form_filler.template_filename).suffix == ".pdf":
                # make sur to have a string value if the form expect a string
                pdf_form_schema = PdfWrapper(template_filepath).schema
                for k, v in fields.items():
                    if k in pdf_form_schema["properties"]:
                        if pdf_form_schema["properties"][k]["type"] == "string":
                            if v is not None:
                                fields[k] = str(v)
                            pass
                if editable:
                    filled = FormWrapper(template_filepath).fill(fields, flatten=False)
                else:
                    filled = PdfWrapper(template_filepath).fill(fields, flatten=False)
                with open(output_filepath, "wb+") as output:
                    output.write(filled.read())
            elif Path(form_filler.template_filename).suffix == ".docx":
                template = DocxTemplate(template_filepath)
                template.render(fields)
                template.save(output_filepath)

            output_filepaths.append(Path(output_filepath))

    return output_filepaths


if __name__ == "__main__":

    fiche = "assets/fiche.xlsx"
    form_fillers = get_form_fillers(fiche)
    # form_fillers = [form_fillers[i] for i in (3,)]
    fill_forms(
        form_fillers,
        template_base_dir="assets/templates/",
    )
